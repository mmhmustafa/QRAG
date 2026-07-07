from sqlalchemy import create_engine, select, func
from sqlalchemy.orm import sessionmaker, selectinload
from pathlib import Path
from app.db import Base
from app.models import Questionnaire, Question, Customer, ProviderConfig, GlobalProviderConfig, AnswerVersion
from app.services import ingest, build_questionnaire, generate_questionnaire, generate_one, config_for, clean_customer_answer, export_xlsx, retrieve
from app.providers import get_llm
from openpyxl import load_workbook

def test_sample_workflow():
    engine=create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    db=sessionmaker(bind=engine,expire_on_commit=False)()
    db.add(GlobalProviderConfig(id=1));customer=Customer(name="Tenant A");db.add(customer);db.flush();db.add(ProviderConfig(customer_id=customer.id));db.commit()
    fixtures=Path(__file__).parent/"fixtures";knowledge=fixtures/"workflow-knowledge.txt";questionnaire=fixtures/"workflow-questions.txt"
    collections=["MedNova Core","Company"]
    ingest(db,knowledge,knowledge.name,customer.id,collections=collections)
    item=build_questionnaire(db,questionnaire,questionnaire.name,customer.id,collections)
    assert item.status=="draft"
    generate_questionnaire(db,item)
    loaded=db.scalar(select(Questionnaire).where(Questionnaire.id==item.id).options(selectinload(Questionnaire.questions).selectinload(Question.answer)))
    answers=[q.answer for q in sorted(loaded.questions,key=lambda x:x.ordinal)]
    assert answers[0].status=="needs_review"
    assert answers[0].classification_reason
    assert answers[0].sources[0]["document"]=="workflow-knowledge.txt"
    assert answers[1].status=="manual_review"
    assert db.scalar(select(func.count(AnswerVersion.id)))==2
    generate_one(db,loaded.questions[0],config_for(db,customer.id),get_llm(config_for(db,customer.id)));db.commit()
    assert db.scalar(select(func.count(AnswerVersion.id)))==3
    answers[0].text="Reviewer edited and approved answer.";answers[0].status="approved";db.commit();stream=export_xlsx(loaded);assert stream.read(2)==b"PK";stream.seek(0);sheet=load_workbook(stream).active
    assert sheet.max_column==2
    # Data starts below the title block (rows 1-2) and styled header (row 4).
    assert sheet.cell(5,2).value=="Reviewer edited and approved answer."
    assert sheet.cell(6,2).value in ("",None)
    internal=export_xlsx(loaded,True);internal_sheet=load_workbook(internal).active
    assert internal_sheet.max_column==12
    assert "MedNova Core" in internal_sheet.cell(5,4).value
    assert "workflow-knowledge.txt" in internal_sheet.cell(5,6).value
    assert internal_sheet.cell(6,2).value=="Manual Review Required"
    answers[0].golden=True;duplicate=Question(customer_id=customer.id,questionnaire_id=item.id,text="Do you encrypt data at rest?",ordinal=99);db.add(duplicate);db.commit()
    reused=generate_one(db,duplicate,config_for(db,customer.id),get_llm(config_for(db,customer.id)));db.commit()
    assert reused.status=="approved_candidate" and reused.reused_from_answer_id==answers[0].id
    cleaned=clean_customer_answer("According to workflow-knowledge.txt [1], MFA is supported. Chunk ID: 17 Similarity 92%",[{"document":"workflow-knowledge.txt"}])
    assert "workflow-knowledge" not in cleaned and "Chunk" not in cleaned and "92%" not in cleaned
    # Retrieval must never mix vectors created by a different embedding model.
    global_config=db.get(GlobalProviderConfig,1);global_config.embedding_model="mock-hash-v2";db.commit()
    assert retrieve(db,"Do you support MFA?",customer.id)==[]
