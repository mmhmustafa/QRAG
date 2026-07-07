"""Customer exports: formatted Excel and questionnaire-order PDF."""
import io
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import select
from app.models import Answer
from app.services import export_xlsx, export_pdf, run_generation, new_progress
from tests.test_generation_progress import make_workspace

def reviewed_workspace(tmp_path):
    db,customer,item=make_workspace(tmp_path,["Do you encrypt data at rest?","Do you support MFA?"])
    run_generation(db,item,new_progress(item.id,customer.id))
    first=db.scalar(select(Answer).where(Answer.customer_id==customer.id).order_by(Answer.id))
    first.status="approved";first.text="Yes, data is encrypted at rest using AES-256.";db.commit()
    db.refresh(item)
    return db,item  # the session must stay alive: exports lazy-load item.questions

def test_xlsx_export_is_formatted(tmp_path):
    db,item=reviewed_workspace(tmp_path)
    ws=load_workbook(export_xlsx(item,False,"ACME Health")).active
    assert ws["A1"].value=="ACME Health" and ws["A1"].font.bold and ws["A1"].font.size==16
    assert "A1:B1" in [str(r) for r in ws.merged_cells.ranges]  # title centered across the table
    assert ws["A4"].value=="Question" and ws["A4"].font.bold and ws["A4"].fill.fgColor.rgb=="FF176B63"
    assert ws.column_dimensions["A"].width==70 and ws.column_dimensions["B"].width==90
    assert ws["A5"].alignment.wrap_text and ws["B5"].alignment.wrap_text
    assert ws.row_dimensions[5].height>0
    assert ws.freeze_panes=="A5"
    rows={ws.cell(row=r,column=1).value:ws.cell(row=r,column=2).value for r in (5,6)}
    assert rows["Do you encrypt data at rest?"]=="Yes, data is encrypted at rest using AES-256."
    assert rows["Do you support MFA?"] in ("",None)  # unapproved answers stay blank in the customer copy

def test_pdf_export_follows_questionnaire_order(tmp_path):
    db,item=reviewed_workspace(tmp_path)
    stream=export_pdf(item,"ACME Health")
    data=stream.read()
    assert data.startswith(b"%PDF")
    text="\n".join(page.extract_text() for page in PdfReader(io.BytesIO(data)).pages)
    assert "ACME Health" in text
    assert "Do you encrypt data at rest?" in text
    assert "Yes, data is encrypted at rest using AES-256." in text
    assert "Do you support MFA?" in text
    assert "(no approved answer)" in text  # unapproved questions keep their slot, clearly unanswered
    # approved answer appears directly after its question
    assert text.index("Do you encrypt data at rest?")<text.index("Yes, data is encrypted at rest using AES-256.")<text.index("Do you support MFA?")
